import json
import openpyxl

wb = openpyxl.load_workbook('/Users/zorantrandafilovic/Downloads/National team - Minhen 2026.xlsx', data_only=True)

# Parse athletes from Paddlers sheet
ps = wb['Paddlers']
race_cols = {}
for col in range(14, 26):  # columns N-Y (14-25 in 1-indexed)
    name = ps.cell(1, col).value
    if name:
        race_cols[col] = name

# Parse athletes
athletes = []
# Determine gender heuristic from race assignments
women_race_names = [n for n in race_cols.values() if 'Women' in n]
open_race_names = [n for n in race_cols.values() if 'open' in n.lower() or 'Open' in n]

for row in range(4, 66):  # rows 4 onwards have athletes
    name = ps.cell(row, 2).value
    if not name or name == 'empty':
        continue
    weight = ps.cell(row, 3).value
    entries = ps.cell(row, 5).value

    assignments = []
    for col, race_name in race_cols.items():
        val = ps.cell(row, col).value
        if val and str(val).strip().lower() == 'x':
            assignments.append(race_name)

    # Determine gender based on race assignments
    in_women = any('Women' in a for a in assignments)
    in_open = any('open' in a.lower() for a in assignments)
    if in_women and not in_open:
        gender = 'F'
    elif in_open and not in_women:
        gender = 'M'
    elif in_women and in_open:
        gender = 'M'  # shouldn't happen
    else:
        # Check from mixed races or guess from name patterns
        # For Serbian names, common female endings
        if name.strip().endswith('ć') or name.strip().endswith('ić'):
            # Could be either - check if weight suggests
            # Use race context from the boat sheets
            gender = 'U'  # unknown, will fix below
        else:
            gender = 'U'

    athletes.append({
        'id': row - 3,
        'name': name.strip(),
        'weight': weight if weight else 0,
        'gender': gender,
        'raceAssignments': assignments
    })

# Fix gender for unknown athletes by checking boat sheets
# Women appearing in Women-only races
women_names = set()
men_names = set()
for a in athletes:
    if any('Women' in r for r in a['raceAssignments']):
        women_names.add(a['name'])
    if any('open' in r.lower() for r in a['raceAssignments']):
        men_names.add(a['name'])

# Also check helm assignments in women's races
for sn in wb.sheetnames:
    if 'Women' in sn and 'TEMPLATE' not in sn:
        ws = wb[sn]
        # Check helm (row 28 in 1-indexed = row 27 in 0-indexed sheet)
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(r, 6).value  # column F = helm name
            if cell_val and cell_val != 'Empty' and cell_val != 'HELM' and cell_val != 'DRUMMER' and cell_val != 'TOTAL':
                women_names.add(cell_val.strip())

# Check all boat layout sheets for names on left/right to determine gender
for sn in wb.sheetnames:
    if 'TEMPLATE' in sn or sn == 'Paddlers' or sn == 'Benches':
        continue
    ws = wb[sn]
    is_women = 'Women' in sn
    is_open = 'open' in sn.lower()
    for r in range(1, ws.max_row + 1):
        for c in [4, 8]:  # columns D and H (left name, right name)
            val = ws.cell(r, c).value
            if val and val not in ('LEFT', 'RIGHT', 'Empty', 'empty', 'reserves:', 'FRONT', 'REAR', 'DRUMMER', 'HELM', 'TOTAL', 'RACE', 'MEDAL', 'Left/Right', 'Top/Down') and not isinstance(val, (int, float)):
                name = val.strip()
                if is_women:
                    women_names.add(name)
                elif is_open:
                    men_names.add(name)

for a in athletes:
    if a['gender'] == 'U':
        if a['name'] in women_names and a['name'] not in men_names:
            a['gender'] = 'F'
        elif a['name'] in men_names:
            a['gender'] = 'M'
        else:
            # Default guess based on Serbian name patterns
            n = a['name'].split()[-1] if a['name'] else ''
            # Female Serbian surnames/names often end in -a (first name)
            fn = a['name'].split()[0] if a['name'] else ''
            if fn.endswith('a'):
                a['gender'] = 'F'
            else:
                a['gender'] = 'M'

# Parse bench factors
bs = wb['Benches']
bench_standard = []
bench_small = []
for col in range(2, 12):
    v = bs.cell(2, col).value
    if v is not None:
        bench_standard.append(float(v))
for col in range(2, 7):
    v = bs.cell(3, col).value
    if v is not None:
        bench_small.append(float(v))

# Parse boat layouts from race sheets
race_sheets = [sn for sn in wb.sheetnames if 'TEMPLATE' not in sn and sn != 'Paddlers' and sn != 'Benches']

def find_athlete_id(name, athletes_list):
    if not name or name in ('Empty', 'empty', ''):
        return None
    name = name.strip()
    for a in athletes_list:
        if a['name'] == name:
            return a['id']
    return None

races = []
layouts = {}

for sn in race_sheets:
    ws = wb[sn]

    # Determine boat type from template reference or row count
    template_ref = ws.cell(2, 1).value or ''
    if 'Small' in template_ref or 'SM ' in sn:
        boat_type = 'small'
        num_rows = 5
    else:
        boat_type = 'standard'
        num_rows = 10

    race_id = sn.replace(' ', '_')

    # Parse distance from name
    distance = ''
    for d in ['200m', '500m', '1000m', '2000m']:
        if d in sn:
            distance = d
            break

    # Parse category
    category = sn.replace(distance, '').strip()

    races.append({
        'id': race_id,
        'name': sn,
        'boatType': boat_type,
        'numRows': num_rows,
        'distance': distance,
        'category': category
    })

    # Parse seating layout
    # Drummer is at row 9 in sheet (1-indexed), name in col 6
    drummer_name = ws.cell(9, 6).value
    drummer_id = find_athlete_id(drummer_name, athletes) if drummer_name and drummer_name != 'Empty' else None

    # Helm
    if boat_type == 'standard':
        helm_row = 38  # row 38 in sheet
    else:
        helm_row = 28
    helm_name = ws.cell(helm_row, 6).value
    helm_id = find_athlete_id(helm_name, athletes) if helm_name and helm_name != 'Empty' else None

    # Seats: left side col D (4), right side col H (8)
    # Weights: left col E (5), right col G (7)
    left = []
    right = []

    # Standard boat: seats at rows 14,16,18,20,22,24,26,28,30,32 (every other row starting from 14)
    # Small boat: seats at rows 14,16,18,20,22 (every other row starting from 14)
    for i in range(num_rows):
        seat_row = 14 + (i * 2)

        left_name = ws.cell(seat_row, 4).value
        right_name = ws.cell(seat_row, 8).value

        left_id = find_athlete_id(left_name, athletes)
        right_id = find_athlete_id(right_name, athletes)

        left.append(left_id)
        right.append(right_id)

    # Reserves
    reserves = []
    if boat_type == 'standard':
        res_rows = [41, 42]  # rows 41-42
    else:
        res_rows = [31]

    for rr in res_rows:
        for c in [4, 8]:
            rn = ws.cell(rr, c).value
            rid = find_athlete_id(rn, athletes)
            if rid:
                reserves.append(rid)

    layouts[race_id] = {
        'drummer': drummer_id,
        'helm': helm_id,
        'left': left,
        'right': right,
        'reserves': reserves
    }

data = {
    'athletes': athletes,
    'benchFactors': {
        'standard': bench_standard,
        'small': bench_small
    },
    'races': races,
    'layouts': layouts
}

with open('src/data/data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Exported {len(athletes)} athletes, {len(races)} races")
print(f"Bench factors - standard: {bench_standard}, small: {bench_small}")
for rid, layout in layouts.items():
    filled = sum(1 for x in layout['left'] + layout['right'] if x is not None)
    print(f"  {rid}: {filled} seats filled, drummer={'yes' if layout['drummer'] else 'no'}, helm={'yes' if layout['helm'] else 'no'}")
